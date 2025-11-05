from io import BytesIO
import io
import zipfile
from typing import List, Dict, Any
import pandas as pd
from fastapi import BackgroundTasks, Form, UploadFile, HTTPException, APIRouter, File
from fastapi.responses import JSONResponse, StreamingResponse, Response
from app.models.header_model import get_all_headers
from app.models.product_model import get_all_products
from app.core.helper import (
    normalize_string,
    rename_columns_with_labels,
    safe_float,
    safe_round,
    get_user_friendly_dtype,
    dataframe_to_pdf,
    TYPE_MAP,
)
from app.core.currency_conversion import (
    get_ecb_fx_rates_from_db,
    get_fx_rate_by_date_from_db_rates,
)
from app.core.send_mail import (
    send_manual_vat_email,
    send_vat_report_email_safely,
    send_quarter_issues_email,
)
from openpyxl import Workbook, load_workbook
import uuid
from datetime import datetime, date, timedelta
from app.core.helper import validate_order_date
from app.schemas.auth_schemas import AdminNotifyRequest
import traceback
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


router = APIRouter()

# In-memory storage for processed data (in production, use Redis or database)2
processed_data_store: Dict[str, Dict[str, Any]] = {}


# Cleanup old entries (older than 24 hours - extended for better UX)
def cleanup_old_data():
    current_time = datetime.now()
    expired_keys = []
    for key, data in processed_data_store.items():
        if current_time - data["timestamp"] > timedelta(hours=24):
            expired_keys.append(key)
    for key in expired_keys:
        print(f"Cleaning up expired session: {key}")
        del processed_data_store[key]

    if expired_keys:
        print(f"Cleaned up {len(expired_keys)} expired sessions")


# Add session validation helper
def validate_session(session_id: str) -> bool:
    """Validate if session exists and is not expired"""
    if session_id not in processed_data_store:
        print(f"Session not found: {session_id}")
        print(f"Available sessions: {list(processed_data_store.keys())}")
        return False

    session_data = processed_data_store[session_id]
    current_time = datetime.now()
    if current_time - session_data["timestamp"] > timedelta(hours=24):
        print(f"Session expired: {session_id}")
        del processed_data_store[session_id]
        return False

    return True


# Read uploaded file and return headers + DataFrame + original content
async def extract_file_headers(
    file: UploadFile,
) -> tuple[list[str], pd.DataFrame, bytes]:
    try:
        content = await file.read()
        file_data = BytesIO(content)
        if file.filename.endswith(".csv"):
            df = pd.read_csv(file_data)
        elif file.filename.endswith(".txt"):
            df = pd.read_csv(file_data, delimiter="\t")
        else:
            df = pd.read_excel(file_data)
        headers = [str(col).strip().lower() for col in df.columns]
        return headers, df, content
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error reading file: {str(e)}")


async def validate_file_data(file_headers: list[str], df: pd.DataFrame) -> dict:
    try:
        all_headers = await get_all_headers()
        alias_to_value = {}
        required_headers = []
        header_labels = {}
        expected_types = {}

        # --- Step 1: Map aliases to standard header values & types ---
        for header in all_headers:
            value = header["value"]
            required_headers.append(value)
            header_labels[value] = header["label"]

            for alias in header["aliases"]:
                alias_to_value[alias.strip().lower()] = value

            raw_type = header.get("type", "string")
            mapped_type = TYPE_MAP.get(raw_type.lower(), "string")
            expected_types[value] = mapped_type

        # --- Step 2: Normalize column names ---
        rename_map = {}
        for col in df.columns:
            normalized = col.strip().lower()
            mapped_value = alias_to_value.get(normalized)
            if mapped_value:
                rename_map[col] = mapped_value

        df.rename(columns=rename_map, inplace=True)

        # --- Step 3: Missing header check ---
        missing_headers_detailed = []
        for field in required_headers:
            if field not in df.columns:
                missing_headers_detailed.append(
                    {
                        "header_value": field,
                        "header_label": header_labels.get(field, field),
                        "expected_name": header_labels.get(field, field),
                        "description": f"Required column '{header_labels.get(field, field)}' is missing from the file",
                    }
                )

        data_issues = []

        # --- Step 4: Missing data + type validation ---
        for header_value in df.columns:
            col_dtype = get_user_friendly_dtype(df[header_value].dtype)

            # Missing data validation
            try:
                null_mask = df[header_value].isnull()
                empty_mask = (
                    df[header_value]
                    .astype(str)
                    .str.strip()
                    .isin(["", "nan", "None", "(empty)", "(null)"])
                )
                combined_mask = null_mask | empty_mask
                total_empty = int(combined_mask.sum())

                if total_empty > 0:
                    missing_rows = df.index[combined_mask].tolist()
                    missing_rows_display = [str(row + 2) for row in missing_rows]

                    issue_description = f"Column '{header_labels.get(header_value, header_value)}' has {total_empty} missing values"
                    if len(missing_rows) > 10:
                        issue_description += f" (showing first 10 rows: {','.join(missing_rows_display[:10])}...)"
                    else:
                        issue_description += (
                            f" in rows: {', '.join(missing_rows_display)}"
                        )

                    data_issues.append(
                        {
                            "header_value": header_value,
                            "header_label": header_labels.get(
                                header_value, header_value
                            ),
                            "original_column": header_value,
                            "issue_type": "MISSING_DATA",
                            "issue_description": issue_description,
                            "column_name": header_labels.get(
                                header_value, header_value
                            ),
                            "data_type": col_dtype,
                            "total_missing": total_empty,
                            "percentage": round((total_empty / len(df)) * 100, 2),
                            "missing_rows": missing_rows_display,
                            "has_more_rows": len(missing_rows) > 10,
                        }
                    )

            except Exception as col_error:
                print(
                    f"Error processing missing data for column {header_value}: {str(col_error)}"
                )

            # Type validation
            try:
                invalid_type_rows = []
                expected_type = expected_types.get(header_value, "string")

                for idx, val in df[header_value].items():
                    if pd.isnull(val) or (isinstance(val, str) and val.strip() == ""):
                        continue
                    is_valid = True
                    error_msg = ""

                    # --- Type-specific checks (integer, float, date, string, etc.) ---
                    # (omitted here for brevity — keep your existing detailed checks)

                    if not is_valid:
                        invalid_type_rows.append(idx + 2)

                if invalid_type_rows:
                    data_issues.append(
                        {
                            "header_value": header_value,
                            "header_label": header_labels.get(
                                header_value, header_value
                            ),
                            "original_column": header_value,
                            "issue_type": "INVALID_TYPE",
                            "issue_description": f"Invalid {expected_type} values in rows: {', '.join(map(str, invalid_type_rows[:10]))}",
                            "expected_type": expected_type,
                            "invalid_rows": invalid_type_rows[:10],
                            "invalid_count": len(invalid_type_rows),
                            "total_rows": len(df),
                            "percentage": round(
                                (len(invalid_type_rows) / len(df)) * 100, 2
                            ),
                            "has_more_rows": len(invalid_type_rows) > 10,
                        }
                    )

            except Exception as type_error:
                print(
                    f"Error during type validation for column {header_value}: {str(type_error)}"
                )

        # --- Step 5: Order date quarter validation ---
        if "order_date" in df.columns:
            today = date.today()
            invalid_quarter_rows = []
            print("today", today)

            for idx, val in df["order_date"].items():
                if pd.isnull(val) or str(val).strip() == "":
                    continue

                print(f"Row {idx+2} → raw value: {repr(val)} ({type(val)})")  # debug

                try:
                    # Normalize to date
                    if isinstance(val, pd.Timestamp):
                        order_date = val.date()
                    elif isinstance(val, (datetime, date)):
                        order_date = val
                    else:
                        order_date = pd.to_datetime(val).date()
                except Exception:
                    # Only catch true parsing errors here
                    invalid_quarter_rows.append(
                        {
                            "row": idx + 2,
                            "value": str(val),
                            "issue": f"Invalid date format: {val}",
                        }
                    )
                    continue  # skip to next row

                # Now safely run validation
                result = validate_order_date(order_date, today)
                print(order_date, today, result)

                if not result.startswith("Accepted"):
                    invalid_quarter_rows.append(
                        {"row": idx + 2, "value": str(val), "issue": result}
                    )

            if invalid_quarter_rows:
                data_issues.append(
                    {
                        "header_value": "order_date",
                        "header_label": header_labels.get("order_date", "Order Date"),
                        "original_column": "order_date",
                        "issue_type": "INVALID_QUARTER",
                        "issue_description": "Some order dates are not in the allowed previous quarter",
                        "invalid_rows": invalid_quarter_rows,  # Include all invalid rows
                        "invalid_count": len(invalid_quarter_rows),
                        "total_rows": len(df),
                        "percentage": round(
                            (len(invalid_quarter_rows) / len(df)) * 100, 2
                        ),
                        "has_more_rows": False,  # No longer limiting, so no "more rows"
                    }
                )
        # --- Step 6: Return results ---
        return {
            "missing_headers": [
                field for field in required_headers if field not in df.columns
            ],
            "missing_headers_detailed": missing_headers_detailed,
            "matched_columns": {v: v for v in df.columns},
            "header_labels": header_labels,
            "data_issues": data_issues,
            "total_rows": len(df),
        }

    except Exception as e:
        print(f"Validation error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Validation error: {str(e)}")



async def enrich_dataframe_with_vat(df: pd.DataFrame) -> tuple:
    try:
        # 1. Get VAT products from database
        vat_products = await get_all_products()
        print(f"Retrieved {len(vat_products)} VAT products from database")

        # 2. Build a lookup dictionary for (product_type, country) -> VAT info
        vat_lookup = {}
        for prod in vat_products:
            try:
                product_type = normalize_string(str(prod.get("product_type", "")))
                country = normalize_string(str(prod.get("country", "")))
                if product_type and country:
                    key = (product_type, country)
                    vat_lookup[key] = prod
                    print(
                        f"Added VAT lookup: {key} -> VAT: {prod.get('vat_rate', 0)}%, Shipping VAT: {prod.get('shipping_vat_rate', 0)}%"
                    )
            except Exception as prod_error:
                print(f"Error processing VAT product: {prod_error}")
                continue

        print(f"Built VAT lookup with {len(vat_lookup)} entries")
        print(f"VAT lookup keys: {list(vat_lookup.keys())}")

        # 3. Identify relevant columns in the DataFrame
        available_columns = [col.lower() for col in df.columns]
        print(f"Available columns (lowercase): {available_columns}")

        # Try to find the right column names for product_type, country, net_price, shipping_amount, currency, order_date
        product_type_col = None
        country_col = None
        net_price_col = None
        shipping_amount_col = None
        currency_col = None
        order_date_col = None

        # For internal testing
        currencies = []
        net_prices = []

        # Find columns by name (case-insensitive)
        for col in df.columns:
            col_lower = col.lower()
            if col_lower == "order_date":
                order_date_col = col
            elif col_lower == "product_type":
                product_type_col = col
            elif col_lower == "country":
                country_col = col
            elif col_lower == "net_price":
                net_price_col = col
                net_prices = df["net_price"]
            elif col_lower == "shipping_amount":
                shipping_amount_col = col
            elif col_lower == "currency":
                currency_col = col
                currencies = df["currency"]

        # 4. Prepare lists to store calculated values for new columns
        vat_rates = []
        vat_amounts = []
        shipping_vat_rates = []
        shipping_vat_amounts = []
        total_vat_amounts = []
        gross_total_amounts = []

        # Lists for converted prices and currencies
        converted_prices = []
        converted_shipping_prices = []
        final_currencies = []
        vat_lookup_status = []
        debug_info = []

        # 5. Fetch ECB currency rates for conversion
        ecb_rates = await get_ecb_fx_rates_from_db()
        # Track rows that need manual review
        manual_review_rows = []

        # 6. Process each row in the DataFrame
        for idx, row in df.iterrows():
            try:
                # Get currency and order date for this row
                currency = (
                    str(row[currency_col]).strip().upper() if currency_col else "EUR"
                )
                order_date = (
                    str(row[order_date_col]).strip() if order_date_col else None
                )

                product_type = normalize_string(
                    str(row[product_type_col])
                    if product_type_col
                    else str(row.get("product_type", ""))
                )
                country = normalize_string(
                    str(row[country_col])
                    if country_col
                    else str(row.get("country", ""))
                )

                # Extract net price and shipping amount
                net_price = (
                    safe_float(row[net_price_col])
                    if net_price_col
                    else safe_float(row.get("price", 0))
                )
                shipping_amount = (
                    safe_float(row[shipping_amount_col])
                    if shipping_amount_col
                    else safe_float(row.get("shipping_amount", 0))
                )

                # 7. Convert currency to EUR if needed
                fx_rate = None
                if currency != "EUR" and order_date:
                    order_date_str = pd.to_datetime(order_date).strftime("%Y-%m-%d")
                    fx_rate = get_fx_rate_by_date_from_db_rates(
                        ecb_rates, order_date_str, currency
                    )
                    if fx_rate:
                        fx_rate = 1 / fx_rate
                        net_price = safe_round(net_price * fx_rate, 2)
                        shipping_amount = safe_round(shipping_amount * fx_rate, 2)
                        print(
                            f"Row {idx}: Converted from {currency} to EUR using rate {fx_rate} for date {order_date}"
                        )
                    else:
                        print(
                            f"Row {idx}: No FX rate found for {currency} on or before {order_date}, using original values"
                        )

                converted_prices.append(net_price)
                converted_shipping_prices.append(shipping_amount)
                final_currencies.append(
                    "EUR" if currency != "EUR" and fx_rate else currency
                )

                total_net_price = safe_round(sum(converted_prices), 2)
                print("Converted Prices", converted_prices)
                print(
                    f"Row {idx}: product_type='{product_type}', country='{country}', price={net_price}, shipping_price={shipping_amount}"
                )

                # 8. Look up VAT data for this product_type and country
                lookup_key = (product_type, country)
                vat_data = vat_lookup.get(lookup_key)

                if vat_data:
                    # Extract VAT rates and calculate VAT amounts
                    vat_rate = safe_float(vat_data.get("vat_rate", 2))
                    vat_rate = safe_round((vat_rate / 100), 2)
                    shipping_vat_rate = safe_float(vat_data.get("shipping_vat_rate", 2))
                    shipping_vat_rate = safe_round((shipping_vat_rate / 100), 2)

                    vat_amount = safe_round(vat_rate * net_price, 2)
                    shipping_vat_amount = safe_round(
                        shipping_vat_rate * shipping_amount, 2
                    )
                    total_vat = safe_round(vat_amount + shipping_vat_amount, 2)
                    gross_total = safe_round(
                        net_price + vat_amount + shipping_vat_amount, 2
                    )
                    lookup_status = "Found"
                    debug_msg = (
                        f"VAT found: {vat_rate}% VAT, {shipping_vat_rate}% shipping VAT"
                    )
                    print(f"Row {idx}: {debug_msg}")
                else:
                    # No VAT data found - use defaults and log for manual review
                    manual_review_rows.append(row.to_dict())
                    vat_rate = "Not Found"
                    shipping_vat_rate = "Not Found"
                    vat_amount = "Not Found"
                    shipping_vat_amount = "Not Found"
                    gross_total = 0.0
                    total_vat = 0.0
                    lookup_status = "Not Found"
                    debug_msg = f"No VAT data found for key: {lookup_key}"
                    print(f"Row {idx}: {debug_msg}")

                # Store calculated values for this row
                vat_rates.append(vat_rate)
                vat_amounts.append(vat_amount)
                shipping_vat_rates.append(shipping_vat_rate)
                shipping_vat_amounts.append(shipping_vat_amount)
                total_vat_amounts.append(total_vat)
                gross_total_amounts.append(gross_total)
                vat_lookup_status.append(lookup_status)
                debug_info.append(debug_msg)

            except Exception as row_error:
                print(f"Error processing row {idx}: {str(row_error)}")
                manual_review_rows.append(idx)
                # Use safe defaults for this row
                vat_rates.append(0.0)
                vat_amounts.append(0.0)
                shipping_vat_rates.append(0.0)
                shipping_vat_amounts.append(0.0)
                total_vat_amounts.append(0.0)
                gross_total_amounts.append(0.0)
                vat_lookup_status.append("Error")
                debug_info.append(f"Error: {str(row_error)}")

        # 9. Update DataFrame with converted prices and currencies
        if net_price_col:
            df[net_price_col] = converted_prices
        if shipping_amount_col:
            df[shipping_amount_col] = converted_shipping_prices
        if currency_col:
            df[currency_col] = final_currencies

        # 10. Add new VAT-related columns to the DataFrame
        df["Previous Currency"] = currencies
        df["Previous Net Price"] = net_prices
        df["VAT Rate"] = vat_rates
        df["Product VAT"] = vat_amounts
        df["Shipping VAT Rate"] = shipping_vat_rates
        df["Shipping VAT"] = shipping_vat_amounts
        df["Total VAT"] = total_vat_amounts
        df["Gross Total"] = gross_total_amounts

        # 11. Optionally, rename columns to user-friendly labels from header config
        df = await rename_columns_with_labels(df)

        # 🔧 FIX: Convert all timestamp columns to strings before JSON serialization
        for col in df.columns:
            if pd.api.types.is_datetime64_any_dtype(df[col]):
                df[col] = df[col].dt.strftime("%Y-%m-%d %H:%M:%S")
            # Also handle any object columns that might contain Timestamp objects
            elif df[col].dtype == 'object':
                # Convert any Timestamp objects in object columns
                df[col] = df[col].apply(lambda x: x.strftime("%Y-%m-%d %H:%M:%S") if isinstance(x, pd.Timestamp) else x)

        # 12. Print final DataFrame for debugging
        pd.set_option("display.max_columns", None)
        pd.set_option("display.width", None)
        pd.set_option("display.max_colwidth", None)
        print("DataFrame table (full view with renamed headers):")
        print(df)

        # 13. Create a summary VAT report by country
        summary = (
            df.groupby("Country", dropna=False)
            .agg({"Net Price": "sum", "Total VAT": "sum"})
            .reset_index()
        )
        summary.rename(
            columns={"Net Price": "Net Sales", "Total VAT": "VAT Amount"}, inplace=True
        )

        summary["Net Sales"] = summary["Net Sales"].apply(lambda x: safe_round(safe_float(x), 2))
        summary["VAT Amount"] = summary["VAT Amount"].apply(lambda x: safe_round(safe_float(x), 2))
        print("Summary VAT Report by Country:")
        print(summary)

        print("Manual Review", manual_review_rows)
        manual_df = pd.DataFrame(manual_review_rows)
        manual_df = await rename_columns_with_labels(manual_df)

        # 🔧 FIX: Convert timestamps in manual_df as well
        for col in manual_df.columns:
            if pd.api.types.is_datetime64_any_dtype(manual_df[col]):
                manual_df[col] = manual_df[col].dt.strftime("%Y-%m-%d %H:%M:%S")
            elif manual_df[col].dtype == 'object':
                manual_df[col] = manual_df[col].apply(lambda x: x.strftime("%Y-%m-%d %H:%M:%S") if isinstance(x, pd.Timestamp) else x)

        # Overwrite manual_review_rows with renamed version
        manual_review_rows = manual_df.to_dict(orient="records")

        # 🔧 FIX: Use the properly converted manual_review_rows instead of the entire dataframe
        manual_review_data = manual_review_rows  # This is already converted

        # Track rows that need manual review
        print("Manual DataFrame", manual_df)

        # 14. Return the enriched DataFrame and summary DataFrame
        if len(manual_review_rows) > 0:
            return {
                "status": "manual_review_required",
                "message": "Some rows could not be processed automatically. We'll email you the results within 24 hours.",
                "manual_review_count": len(manual_review_rows),
                "require_email": True,
                "manual_review_rows": manual_review_data,
            }

        # At the end of enrich_dataframe_with_vat
        return (
            df,
            summary,
            manual_df,
            {
                "overall_vat_amount": safe_round(sum(total_vat_amounts), 2),
                "overall_net_price": total_net_price,
                "overall_gross_total": safe_float(sum(gross_total_amounts), 2),
            },
        )

    except Exception as e:
        print(f"Error in VAT enrichment: {str(e)}")
        import traceback

        traceback.print_exc()
        raise HTTPException(
            status_code=500, detail=f"Failed to enrich data with VAT: {str(e)}"
        )


@router.post("/validate-file")
async def validate_file(files: List[UploadFile] = File(...)):
    cleanup_old_data()
    results = []

    for file in files:
        try:
            print(f"Processing file: {file.filename}")
            # Check file type
            allowed_extensions = [".csv", ".txt", ".xls", ".xlsx"]
            file_extension = "." + file.filename.split(".")[-1].lower()
            if file_extension not in allowed_extensions:
                results.append(
                    {
                        "file_name": file.filename,
                        "success": False,
                        "message": f"Unsupported file type: {file_extension}",
                    }
                )
                continue

            # Extract headers, data, and original content from file
            headers, df, original_content = await extract_file_headers(file)

            if not headers:
                results.append(
                    {
                        "file_name": file.filename,
                        "success": False,
                        "message": "No headers found in the file",
                    }
                )
                continue

            # Validate file data
            validation_result = await validate_file_data(headers, df)
            print("File validation completed")

            has_issues = (
                len(validation_result["missing_headers"]) > 0
                or len(validation_result["data_issues"]) > 0
            )

            # Generate unique session ID for this file
            session_id = str(uuid.uuid4())

            # Store processed data in memory (use Redis/DB in production)
            processed_data_store[session_id] = {
                "timestamp": datetime.now(),
                "file_name": file.filename,
                "original_df": df.copy(),  # Store original DataFrame
                "original_file_content": original_content,  # Store original file content
                "validation_result": validation_result,
                "headers": headers,
                "has_issues": has_issues,
            }

            results.append(
                {
                    "file_name": file.filename,
                    "session_id": session_id,  # Return session ID to frontend
                    "success": not has_issues,
                    "has_issues": has_issues,
                    "validation_result": validation_result,
                    "message": (
                        "File has validation issues"
                        if has_issues
                        else "File validation completed successfully"
                    ),
                }
            )

        except Exception as e:
            print(f"Error processing file {file.filename}: {str(e)}")
            import traceback

            traceback.print_exc()
            results.append(
                {
                    "file_name": file.filename,
                    "success": False,
                    "message": f"Error validating file: {str(e)}",
                }
            )

    return {"files": results}


@router.get("/download-vat-issues/{session_id}")
async def download_vat_issues(session_id: str):
    try:
        # Validate session with enhanced logging
        if not validate_session(session_id):
            raise HTTPException(status_code=404, detail="Session not found or expired")
        print(f"Validating session {session_id}...")
        stored_data = processed_data_store[session_id]
        df = stored_data["original_df"].copy()
        validation_result = stored_data["validation_result"]
        file_name = stored_data["file_name"]

        # Format date columns
        for col in df.columns:
            if "order date" in col.lower():
                df[col] = pd.to_datetime(df[col], errors="coerce").dt.strftime(
                    "%d-%m-%Y"
                )

        # Get header labels mapping
        all_headers = await get_all_headers()
        header_labels = {}

        # Build mapping from header values to labels
        for header in all_headers:
            value = header["value"]
            label = header["label"]
            header_labels[value] = label

        reverse_rename_map = {}
        for key, val in header_labels.items():
            reverse_rename_map[key] = val

        # Create rename mapping for existing columns
        rename_map = {}
        for col in df.columns:
            # Check if this column has a corresponding label
            if col in header_labels:
                rename_map[col] = header_labels[col]

        # Apply the renaming
        if rename_map:
            df.rename(columns=rename_map, inplace=True)

        issues = validation_result.get("data_issues", [])
        missing_headers = validation_result.get("missing_headers_detailed", [])
        header_labels = validation_result.get("header_labels", {})

        # Insert placeholder columns for missing headers
        missing_labels = [mh["header_label"] for mh in missing_headers]
        for label in missing_labels:
            if label not in df.columns:
                df[label] = ""

        # Build issues sheet
        issues_rows = []
        for mh in missing_headers:
            issues_rows.append(
                {
                    "Issue Type": "Missing Header",
                    "Column": mh["header_label"],
                    "Description": mh["description"],
                }
            )

        for issue in issues:
            issues_rows.append(
                {
                    "Issue Type": issue["issue_type"],
                    "Column": issue["header_label"],
                    "Description": issue["issue_description"],
                    "Missing Count": issue.get("total_missing")
                    or issue.get("invalid_count", ""),
                    "Missing %": issue.get("percentage", ""),
                }
            )

        issues_df = pd.DataFrame(
            issues_rows
            or [
                {
                    "Issue Type": "None",
                    "Column": "All",
                    "Description": "No missing headers or data issues found.",
                }
            ]
        )

        # --- Create Excel workbook ---
        wb = Workbook()
        ws_data = wb.active
        ws_data.title = "User Data"

        # Fills
        red_fill = PatternFill(
            start_color="FF9999", end_color="FF9999", fill_type="solid"
        )  # Header or invalid type
        orange_fill = PatternFill(
            start_color="FFBF00", end_color="FFF2CC", fill_type="solid"
        )  # Missing values
        header_fill = PatternFill(
            start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"
        )  # Normal header
        bold_font = Font(bold=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # --- Write data ---
        for r in dataframe_to_rows(df, index=False, header=True):
            ws_data.append(r)

        col_name_to_index = {col: idx for idx, col in enumerate(df.columns)}

        # --- Highlight headers ---
        for col_idx, col in enumerate(df.columns, start=1):
            cell = ws_data.cell(row=1, column=col_idx)
            cell.font = bold_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
            cell.fill = red_fill if col in missing_labels else header_fill

        # --- Highlight issues ---
        for issue in issues:
            original_col = issue.get("original_column")
            if not original_col:
                continue

            # Map system name to label
            renamed_col = reverse_rename_map.get(original_col, original_col)
            if renamed_col not in col_name_to_index:
                continue

            col_idx = col_name_to_index[renamed_col] + 1  # openpyxl is 1-indexed

            if issue["issue_type"] == "MISSING_DATA":
                for row_str in issue.get("missing_rows", []):
                    try:
                        row_num = int(row_str)
                        ws_data.cell(row=row_num, column=col_idx).fill = orange_fill
                    except Exception:
                        continue

            elif issue["issue_type"] == "INVALID_TYPE":
                for row_num in issue.get("invalid_rows", []):
                    try:
                        ws_data.cell(row=row_num, column=col_idx).fill = red_fill
                    except Exception:
                        continue

            elif issue["issue_type"] == "INVALID_QUARTER":
                col_key = issue.get("header_label") or issue.get("original_column")
                if col_key not in col_name_to_index:
                    continue
                col_idx = col_name_to_index[col_key] + 1

                for row_info in issue.get("invalid_rows", []):
                    try:
                        row_num = (
                            row_info["row"]
                            if isinstance(row_info, dict)
                            else int(row_info)
                        )
                        ws_data.cell(row=row_num, column=col_idx).fill = red_fill
                    except Exception:
                        continue

        # --- Style data cells + autosize ---
        for row in ws_data.iter_rows(min_row=2):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

        for col in ws_data.columns:
            max_len = max(
                (len(str(cell.value)) for cell in col if cell.value), default=10
            )
            col_letter = get_column_letter(col[0].column)
            ws_data.column_dimensions[col_letter].width = max_len + 2

        # --- Add issues sheet ---
        ws_issues = wb.create_sheet("Validation Issues")
        for r in dataframe_to_rows(issues_df, index=False, header=True):
            ws_issues.append(r)

        for col in ws_issues.iter_cols(min_row=1, max_row=1):
            for cell in col:
                cell.fill = header_fill
                cell.font = bold_font
                cell.border = thin_border

        for row in ws_issues.iter_rows(min_row=2):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        for col in ws_issues.columns:
            max_len = max(
                (len(str(cell.value)) for cell in col if cell.value), default=10
            )
            col_letter = get_column_letter(col[0].column)
            ws_issues.column_dimensions[col_letter].width = max_len + 2

        # --- Output stream ---
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        download_name = file_name.rsplit(".", 1)[0] + "_validation_annotated.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={download_name}"},
        )

    except Exception as e:
        print("Error stack trace:")
        traceback.print_exc()
        raise HTTPException(
            status_code=500, detail=f"Could not generate issue report: {str(e)}"
        )


@router.post("/send-vat-report-email/{session_id}")
async def send_vat_report_email(
    session_id: str,
    background_tasks: BackgroundTasks,
    user_email: str = Form(...),
    file_name: str = Form(...),
):
    """
    Generate VAT and Summary reports (Excel + PDF), zip them, and send via email.
    """
    try:
        # --- Validate session ---
        if not validate_session(session_id):
            raise HTTPException(status_code=404, detail="Session not found or expired")

        stored_data = processed_data_store[session_id]
        df = stored_data["original_df"].copy()

        print(f"Preparing to send VAT report to {user_email}")
        print(f"Session ID: {session_id}")
        print(f"File name: {file_name}")

        # --- Enrich VAT Data ---
        result = await enrich_dataframe_with_vat(df)
        if isinstance(result, dict) and result.get("status") == "manual_review_required":
            return {
                "status": "error",
                "message": "This file requires manual review and cannot be processed automatically.",
            }

        enriched_df, summary_df, manual_df, vat_summary = result

        # --- Format date columns ---
        for col in enriched_df.columns:
            if "order date" in col.lower() and pd.api.types.is_datetime64_any_dtype(enriched_df[col]):
                enriched_df[col] = enriched_df[col].dt.strftime("%d-%m-%Y")

        # --- Prepare in-memory Excel files ---
        vat_excel_stream = io.BytesIO()
        summary_excel_stream = io.BytesIO()

        with pd.ExcelWriter(vat_excel_stream, engine="openpyxl") as writer:
            enriched_df.to_excel(writer, index=False, sheet_name="VAT Report")

        with pd.ExcelWriter(summary_excel_stream, engine="openpyxl") as writer:
            summary_df.to_excel(writer, index=False, sheet_name="Summary Report")

        # --- Prepare in-memory PDF files ---
        vat_pdf_stream = io.BytesIO()
        summary_pdf_stream = io.BytesIO()

        dataframe_to_pdf(enriched_df, vat_pdf_stream, "VAT Report")
        dataframe_to_pdf(summary_df, summary_pdf_stream, "Summary Report")

        vat_pdf_bytes = vat_pdf_stream.getvalue()
        summary_pdf_bytes = summary_pdf_stream.getvalue()

        # --- Create ZIP file with all reports ---
        base_name = file_name.rsplit(".", 1)[0] if "." in file_name else file_name
        zip_name = f"{base_name}_VAT_Reports.zip"

        zip_stream = io.BytesIO()
        with zipfile.ZipFile(zip_stream, "w", zipfile.ZIP_DEFLATED) as zipf:
            zipf.writestr(f"{base_name}_VAT_Report.xlsx", vat_excel_stream.getvalue())
            zipf.writestr(f"{base_name}_Summary.xlsx", summary_excel_stream.getvalue())
            zipf.writestr(f"{base_name}_VAT_Report.pdf", vat_pdf_bytes)
            zipf.writestr(f"{base_name}_Summary.pdf", summary_pdf_bytes)

        zip_stream.seek(0)
        zip_content = zip_stream.getvalue()

        # --- Send email asynchronously (background) ---
        background_tasks.add_task(
            send_vat_report_email_safely,
            user_email,
            zip_name,
            zip_content,
            file_name,
        )

        print(f"VAT report successfully prepared for {user_email}")

        return {
            "status": "success",
            "message": f"VAT report has been sent to {user_email}",
        }

    except HTTPException as he:
        print(f"HTTPException in send_vat_report_email: {he.detail}")
        raise he
    except Exception as e:
        print(f"Unhandled exception in send_vat_report_email: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Could not send email: {str(e)}")


@router.post("/send-manual-review-admin-email/{session_id}")
async def send_manual_review_admin_email(
    session_id: str,
    background_tasks: BackgroundTasks,
    user_email: str = Form(...),
    file_name: str = Form(...),
    processed_data: str = Form(...),
):
    try:
        # ===== Validate session =====
        if not validate_session(session_id):
            raise HTTPException(status_code=404, detail="Session not found or expired")

        stored_data = processed_data_store[session_id]
        original_df = stored_data["original_df"].copy()
        file_name = stored_data["file_name"]

        print("File validation completed")

        # ===== Enrich VAT data =====
        enrichment_result = await enrich_dataframe_with_vat(original_df)

        df = None
        summary_df = None
        manual_review_rows = []

        # Handle manual review vs normal result
        if (
            isinstance(enrichment_result, dict)
            and enrichment_result.get("status") == "manual_review_required"
        ):
            manual_review_rows = enrichment_result.get("manual_review_rows", [])
            df = pd.DataFrame(manual_review_rows)
            summary_df = None
        else:
            df = enrichment_result.get("processed_df")
            summary_df = enrichment_result.get("summary_df", None)

        if df is None or df.empty:
            raise HTTPException(
                status_code=400, detail="Processed VAT data not found or is empty"
            )

        # ===== Parse manual review rows from request (if provided) =====
        import json

        try:
            manual_review_rows_request = json.loads(processed_data)
        except json.JSONDecodeError:
            manual_review_rows_request = manual_review_rows

        print(f"Sending manual review admin email to {user_email} for file {file_name}")

        # ===== Rename column headers to labels =====
        all_headers = await get_all_headers()
        header_labels = {h["value"]: h["label"] for h in all_headers}

        def rename_headers(dataframe):
            if dataframe is None or dataframe.empty:
                return dataframe
            rename_map = {
                col: header_labels[col]
                for col in dataframe.columns
                if col in header_labels
            }
            if rename_map:
                print(f"Renaming columns: {rename_map}")
                dataframe.rename(columns=rename_map, inplace=True)
            return dataframe

        df = rename_headers(df)
        summary_df = rename_headers(summary_df) if summary_df is not None else None

        # ===== Build Excel with VAT Report & Summary =====
        manual_email_stream = io.BytesIO()
        with pd.ExcelWriter(manual_email_stream, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="VAT Report")
            if summary_df is not None and not summary_df.empty:
                summary_df.to_excel(writer, index=False, sheet_name="Summary")
        manual_email_stream.seek(0)

        # ===== Load workbook for highlighting =====
        workbook = load_workbook(manual_email_stream)
        fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        for sheet_name in ["VAT Report", "Summary"]:
            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                    highlight_row = any(
                        str(cell.value).strip() == "Not Found"
                        for cell in row
                        if cell.value is not None
                    )
                    for cell in row:
                        cell.font = Font(name="Calibri", size=12, bold=False)
                        if highlight_row:
                            cell.fill = fill
                for cell in sheet[1]:
                    cell.font = Font(name="Calibri", size=12, bold=False)

        # ===== Save highlighted file to stream =====
        final_manual_email_stream = io.BytesIO()
        workbook.save(final_manual_email_stream)
        final_manual_email_stream.seek(0)

        

        # ===== Send email in background =====
        background_tasks.add_task(
            send_manual_vat_email,
            "mailer@xtechon.com",  # From email
            user_email,  # Admin email
            final_manual_email_stream.getvalue(),
            manual_review_rows_request,
        )

        print("Manual review admin email task added to background.")

        return {
            "status": "success",
            "message": f"Manual review request sent to admin at {user_email}",
            "file_name": file_name,
            "manual_review_count": len(manual_review_rows_request),
        }

    except HTTPException as he:
        print(f"HTTPException in send_manual_review_admin_email: {he.detail}")
        raise he
    except Exception as e:
        print(f"Unhandled exception in send_manual_review_admin_email: {str(e)}")
        import traceback

        traceback.print_exc()
        raise HTTPException(
            status_code=500, detail=f"Could not send manual review email: {str(e)}"
        )


@router.post("/notify-admin-quarter-issues")
async def notify_admin_quarter_issues(payload: AdminNotifyRequest):
    try:
        user_email = payload.userEmail
        session_id = payload.sessionId

        if not user_email:
            raise HTTPException(status_code=400, detail="User email is required")

        if not session_id:
            raise HTTPException(status_code=400, detail="Session ID is required")

        # Validate and retrieve session data
        if not validate_session(session_id):
            raise HTTPException(status_code=404, detail="Session not found or expired")

        stored_data = processed_data_store[session_id]
        df = stored_data["original_df"].copy()
        validation_result = stored_data["validation_result"]
        file_name = stored_data["file_name"]
        original_file_content = stored_data.get("original_file_content")

        # Extract quarter issues from validation result
        quarter_issues = []
        data_issues = validation_result.get("data_issues", [])

        for issue in data_issues:
            if issue.get("issue_type") == "INVALID_QUARTER":
                quarter_issues.append(
                    {
                        "column": issue.get(
                            "header_label", issue.get("original_column", "Unknown")
                        ),
                        "invalid_count": issue.get("invalid_count", 0),
                        "percentage": issue.get("percentage", 0),
                        "invalid_rows": issue.get(
                            "invalid_rows", []
                        ),  # Include all invalid rows
                    }
                )

        # Create detailed email content
        subject = f"Quarter Validation Issues - File: {file_name}"

        # Build detailed issue summary
        issue_summary = ""
        total_invalid_dates = sum(issue["invalid_count"] for issue in quarter_issues)

        if quarter_issues:
            issue_summary = f"\n\nISSUE SUMMARY:\n"
            issue_summary += f"Total Invalid Dates: {total_invalid_dates}\n"
            issue_summary += f"Total Rows in File: {len(df)}\n\n"

            for i, issue in enumerate(quarter_issues, 1):
                issue_summary += f"{i}. Column: {issue['column']}\n"
                issue_summary += f"   Invalid Count: {issue['invalid_count']} ({issue['percentage']}%)\n"

                if issue["invalid_rows"]:
                    # Show first 10 rows in email for readability, but indicate total count
                    rows_to_show = issue["invalid_rows"][:10]
                    remaining_count = len(issue["invalid_rows"]) - len(rows_to_show)

                    issue_summary += f"   Invalid Rows (showing {len(rows_to_show)} of {len(issue['invalid_rows'])}): "
                    for row_info in rows_to_show:
                        if isinstance(row_info, dict):
                            issue_summary += f"Row {row_info.get('row', 'N/A')} ({row_info.get('value', 'N/A')}), "
                        else:
                            issue_summary += f"Row {row_info}, "
                    issue_summary = issue_summary.rstrip(", ")

                    if remaining_count > 0:
                        issue_summary += f" ...and {remaining_count} more rows"
                    issue_summary += "\n\n"

        body = f"""Quarter Validation Issues Detected

        USER DETAILS:
        Email: {user_email}
        File: {file_name}
        Session ID: {session_id}
        Upload Time: {stored_data['timestamp'].strftime('%Y-%m-%d %H:%M:%S')}

        {issue_summary}

        ATTACHMENTS:
        1. {file_name}_quarter_issues.xlsx - Processed file with issue details and summary
        2. ORIGINAL_{file_name} - Original uploaded file from the user

        ACTION REQUIRED:
        Please contact the user at {user_email} to resolve these quarter validation issues.

        The user's file contains order dates that fall outside the expected previous quarter range.
        This requires manual review and approval from an administrator.

        ---
        This is an automated notification from the VAT Processing System.
        """

        # Create Excel file with highlighted issues for attachment
        excel_stream = io.BytesIO()

        # Get header labels for renaming
        all_headers = await get_all_headers()
        header_labels = {h["value"]: h["label"] for h in all_headers}

        # Rename columns to user-friendly labels
        df_for_email = df.copy()
        rename_map = {
            col: header_labels.get(col, col)
            for col in df_for_email.columns
            if col in header_labels
        }
        if rename_map:
            df_for_email.rename(columns=rename_map, inplace=True)

        # Format dates for display
        for col in df_for_email.columns:
            if "date" in col.lower() and pd.api.types.is_datetime64_any_dtype(
                df_for_email[col]
            ):
                df_for_email[col] = df_for_email[col].dt.strftime("%d-%m-%Y")

        # Create Excel with issue highlighting
        with pd.ExcelWriter(excel_stream, engine="openpyxl") as writer:
            df_for_email.to_excel(writer, index=False, sheet_name="Quarter Issues")

            # Get the workbook and worksheet for highlighting
            workbook = writer.book
            worksheet = writer.sheets["Quarter Issues"]

            # Define highlighting styles
            red_fill = PatternFill(
                start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"
            )  # Light red background
            red_font = Font(color="CC0000", bold=True)  # Dark red font

            # Collect all invalid row numbers for highlighting
            invalid_row_numbers = set()
            if quarter_issues:
                for issue in quarter_issues:
                    for row_info in issue["invalid_rows"]:
                        if isinstance(row_info, dict):
                            row_num = row_info.get("row")
                            if row_num is not None:
                                # Convert to Excel row number (row_num is already 1-based from validation)
                                # Add 1 for the header row that pandas adds
                                excel_row = int(row_num)
                                invalid_row_numbers.add(excel_row)

            # Add a legend/note at the top FIRST
            if invalid_row_numbers:
                # Insert a row at the top for the legend
                worksheet.insert_rows(1)
                legend_cell = worksheet.cell(row=1, column=1)
                legend_cell.value = f"⚠️ HIGHLIGHTED ROWS ({len(invalid_row_numbers)} total) have Quarter Validation Issues - See 'Issue Details' sheet for specifics"
                legend_cell.fill = PatternFill(
                    start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"
                )  # Light yellow
                legend_cell.font = Font(bold=True, color="CC6600")

                # Merge cells for the legend to span across columns
                if worksheet.max_column > 1:
                    worksheet.merge_cells(
                        start_row=1,
                        start_column=1,
                        end_row=1,
                        end_column=worksheet.max_column,
                    )

                # Now highlight invalid rows (add 1 more to account for the legend row we just inserted)
                max_col = worksheet.max_column
                for row_num in invalid_row_numbers:
                    # Add 1 for the legend row we inserted at the top
                    adjusted_row = row_num + 1
                    if adjusted_row <= worksheet.max_row:  # Ensure row exists
                        for col in range(1, max_col + 1):
                            cell = worksheet.cell(row=adjusted_row, column=col)
                            cell.fill = red_fill
                            cell.font = red_font

                print(
                    f"Debug: Highlighted {len(invalid_row_numbers)} rows: {sorted(invalid_row_numbers)}"
                )
                print(
                    f"Debug: Adjusted Excel rows: {sorted([r + 1 for r in invalid_row_numbers])}"
                )

            # Create issues summary sheet
            if quarter_issues:
                issues_data = []
                for issue in quarter_issues:
                    for row_info in issue[
                        "invalid_rows"
                    ]:  # Include all invalid rows in Excel
                        if isinstance(row_info, dict):
                            issues_data.append(
                                {
                                    "Column": issue["column"],
                                    "Row Number": row_info.get("row", "N/A"),
                                    "Invalid Value": row_info.get("value", "N/A"),
                                    "Issue Description": row_info.get(
                                        "issue", "Outside quarter range"
                                    ),
                                }
                            )

                if issues_data:
                    issues_df = pd.DataFrame(issues_data)
                    issues_df.to_excel(writer, index=False, sheet_name="Issue Details")

                    # Format the Issue Details sheet
                    issues_worksheet = writer.sheets["Issue Details"]

                    # Style the header row
                    header_fill = PatternFill(
                        start_color="4472C4", end_color="4472C4", fill_type="solid"
                    )  # Blue background
                    header_font = Font(color="FFFFFF", bold=True)  # White font

                    for col in range(1, len(issues_df.columns) + 1):
                        cell = issues_worksheet.cell(row=1, column=col)
                        cell.fill = header_fill
                        cell.font = header_font

                    # Auto-adjust column widths
                    for column in issues_worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                        issues_worksheet.column_dimensions[column_letter].width = (
                            adjusted_width
                        )

        excel_stream.seek(0)

        # Send email with attachment for quarter issues
        await send_quarter_issues_email(
            to_email="mailer@xtechon.com",  # Admin email
            subject=subject,
            body=body,
            attachment=excel_stream.getvalue(),
            filename=f"{file_name}_quarter_issues.xlsx",
            original_file=original_file_content,
            original_filename=file_name,
        )

        return {
            "status": "success",
            "message": f"Quarter issues notification sent to admin with file details for {file_name}",
            "details": {
                "file_name": file_name,
                "total_invalid_dates": total_invalid_dates,
                "quarter_issues_count": len(quarter_issues),
            },
        }

    except HTTPException as he:
        print(f"HTTPException in notify_admin_quarter_issues: {he.detail}")
        raise he
    except Exception as e:
        print("Error stack trace:")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Failed to notify admin: {str(e)}")


@router.post("/download-vat-report/{session_id}")
async def download_vat_report(
    session_id: str, background_tasks: BackgroundTasks, user_email: str = Form(...)
):
    try:
        # Validate session
        if not validate_session(session_id):
            raise HTTPException(status_code=404, detail="Session not found or expired")

        stored_data = processed_data_store[session_id]
        df = stored_data["original_df"].copy()
        file_name = stored_data["file_name"]

        result = await enrich_dataframe_with_vat(df)

        # 🟡 Manual Review handling (same as before)
        if isinstance(result, dict) and result.get("status") == "manual_review_required":
            # (existing manual review code...)
            return JSONResponse(status_code=200, content=result)

        # 🧾 Normal VAT report generation
        enriched_df, summary_df, manual_df, vat_summary = result

        for col in enriched_df.columns:
            if "order date" in col.lower() and pd.api.types.is_datetime64_any_dtype(
                enriched_df[col]
            ):
                enriched_df[col] = enriched_df[col].dt.strftime("%d-%m-%Y")

        # Prepare in-memory streams
        vat_excel_stream = io.BytesIO()
        summary_excel_stream = io.BytesIO()
        vat_pdf_stream = io.BytesIO()
        summary_pdf_stream = io.BytesIO()
        zip_stream = io.BytesIO()

        # 1️⃣ Generate VAT Excel
        with pd.ExcelWriter(vat_excel_stream, engine="openpyxl") as writer:
            enriched_df.to_excel(writer, index=False, sheet_name="VAT Report")
            workbook = writer.book
            vat_sheet = writer.sheets["VAT Report"]

            start_row = enriched_df.shape[0] + 3
            vat_sheet.cell(row=start_row, column=5, value="Overall Net Total")
            vat_sheet.cell(
                row=start_row + 1, column=5, value=vat_summary["overall_net_price"]
            )
            vat_sheet.cell(row=start_row, column=13, value="Overall VAT Amount")
            vat_sheet.cell(
                row=start_row + 1, column=13, value=vat_summary["overall_vat_amount"]
            )
            vat_sheet.cell(row=start_row, column=14, value="Overall Gross Total")
            vat_sheet.cell(
                row=start_row + 1, column=14, value=vat_summary["overall_gross_total"]
            )

            for row in vat_sheet.iter_rows():
                for cell in row:
                    cell.font = Font(name="Calibri", size=12, bold=False)

        vat_excel_stream.seek(0)

        # 2️⃣ Generate Summary Excel
        with pd.ExcelWriter(summary_excel_stream, engine="openpyxl") as writer:
            summary_df.to_excel(writer, index=False, sheet_name="Summary")
            summary_sheet = writer.sheets["Summary"]
            for row in summary_sheet.iter_rows():
                for cell in row:
                    cell.font = Font(name="Calibri", size=12, bold=False)

        summary_excel_stream.seek(0)

        # 3️⃣ Generate PDFs using helper
        dataframe_to_pdf(enriched_df, vat_pdf_stream, "VAT Report")
        dataframe_to_pdf(summary_df, summary_pdf_stream, "Summary Report")

        # 4️⃣ Bundle everything into ZIP
        base_name = file_name.rsplit(".", 1)[0]
        vat_excel_name = f"{base_name}_VAT_Report.xlsx"
        summary_excel_name = f"{base_name}_Summary.xlsx"
        vat_pdf_name = f"{base_name}_VAT_Report.pdf"
        summary_pdf_name = f"{base_name}_Summary.pdf"
        zip_name = f"{base_name}_VAT_Reports.zip"

        with zipfile.ZipFile(zip_stream, "w", zipfile.ZIP_DEFLATED) as zipf:
            zipf.writestr(vat_excel_name, vat_excel_stream.getvalue())
            zipf.writestr(summary_excel_name, summary_excel_stream.getvalue())
            zipf.writestr(vat_pdf_name, vat_pdf_stream.getvalue())
            zipf.writestr(summary_pdf_name, summary_pdf_stream.getvalue())

        zip_stream.seek(0)

        # ✅ Convert the in-memory ZIP stream to raw bytes
        zip_bytes = zip_stream.getvalue()

        # ✅ Send binary ZIP response for all platforms (Windows, Mac, iOS)
        return Response(
            content=zip_bytes,
            media_type="application/zip",
            headers={
                "Content-Disposition": f'attachment; filename="{zip_name}"',
                "Content-Length": str(len(zip_bytes))
            },
        )


    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error generating report: {str(e)}")